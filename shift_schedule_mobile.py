# shift_schedule_mobile.py - Mobile Optimized Version
import streamlit as st
import pandas as pd
import calendar
import plotly.express as px
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
import streamlit.components.v1 as components
from datetime import datetime

# ---------------- PAGE SETUP (MOBILE OPTIMIZED) ----------------
st.set_page_config(
    layout="centered",  # Better for mobile
    page_title="Shift Schedule",
    page_icon="📅",
    initial_sidebar_state="collapsed"  # Save space on mobile
)

# ---------------- MOBILE CSS ----------------
st.markdown("""
    <style>
    /* Main container padding for mobile */
    .main > div {
        padding: 0 0.5rem;
    }
    
    /* Make all buttons touch-friendly */
    .stButton button, .stDownloadButton button {
        min-height: 48px;
        width: 100%;
        margin: 5px 0;
        border-radius: 12px;
        font-size: 16px;
    }
    
    /* Improve select boxes and inputs for touch */
    .stSelectbox, .stNumberInput, .stTextInput, .stDateInput {
        min-height: 48px;
    }
    
    div[data-baseweb="select"] > div {
        min-height: 48px;
    }
    
    /* Mobile-optimized data editor */
    .stDataEditor {
        font-size: 12px;
        overflow-x: auto;
    }
    
    .stDataEditor [data-testid="stDataEditorResizable"] {
        min-width: 100%;
    }
    
    /* Better cards and containers */
    .element-container, .stAlert, .stInfo, .stSuccess, .stWarning {
        margin: 8px 0;
        border-radius: 12px;
    }
    
    /* Metrics for mobile */
    [data-testid="stMetricValue"] {
        font-size: 24px;
    }
    
    [data-testid="stMetricLabel"] {
        font-size: 12px;
    }
    
    /* Tabs for mobile (if used) */
    .stTabs [data-baseweb="tab-list"] {
        gap: 4px;
        flex-wrap: wrap;
    }
    
    .stTabs [data-baseweb="tab"] {
        min-height: 44px;
        padding: 8px 12px;
        font-size: 13px;
    }
    
    /* Sidebar improvements */
    section[data-testid="stSidebar"] {
        width: 85vw !important;
    }
    
    /* Column adjustments for mobile */
    @media (max-width: 768px) {
        .row-widget.stHorizontal {
            flex-direction: column;
        }
        
        .row-widget.stHorizontal > div {
            width: 100% !important;
            margin: 4px 0;
        }
        
        /* Make dataframe responsive */
        .dataframe {
            font-size: 11px;
        }
        
        .dataframe td, .dataframe th {
            padding: 6px 4px;
        }
    }
    
    /* Shift legend cards for mobile */
    .shift-legend {
        background: #f8f9fa;
        border-radius: 12px;
        padding: 8px;
        margin: 4px;
        text-align: center;
        font-size: 12px;
    }
    
    /* Touch highlight */
    button:active {
        transform: scale(0.98);
        transition: transform 0.05s;
    }
    
    /* Hide extra whitespace */
    .block-container {
        padding-top: 1rem;
        padding-bottom: 0rem;
    }
    </style>
""", unsafe_allow_html=True)

# ---------------- PWA SETUP ----------------
# Add PWA meta tags for installability
components.html("""
    <link rel="manifest" href="/manifest.json">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
    <meta name="apple-mobile-web-app-title" content="Shift Schedule">
    <link rel="apple-touch-icon" href="/icon-192.png">
    <meta name="theme-color" content="#1976D2">
    
    <script>
        // Register Service Worker
        if ('serviceWorker' in navigator) {
            window.addEventListener('load', () => {
                navigator.serviceWorker.register('/service-worker.js')
                    .then(reg => console.log('Service Worker registered'))
                    .catch(err => console.log('Service Worker failed:', err));
            });
        }
        
        // PWA Installation
        let deferredPrompt;
        window.addEventListener('beforeinstallprompt', (e) => {
            e.preventDefault();
            deferredPrompt = e;
            // Show install button if needed
            window.parent.postMessage({type: 'pwa-ready'}, '*');
        });
        
        function installPWA() {
            if (deferredPrompt) {
                deferredPrompt.prompt();
                deferredPrompt.userChoice.then(choice => {
                    if (choice.outcome === 'accepted') {
                        console.log('PWA installed');
                    }
                    deferredPrompt = null;
                });
            }
        }
    </script>
""", height=0)

# ---------------- TITLE WITH INSTALL BUTTON ----------------
col_title, col_install = st.columns([4, 1])
with col_title:
    st.title("📅 Shift Schedule")
    st.caption("SBD-ELE | Mobile Optimized")
with col_install:
    # Placeholder for install button
    pass

# ---------------- SESSION INITIALIZATION ----------------
if "employees" not in st.session_state:
    st.session_state.employees = []

if "schedule" not in st.session_state:
    st.session_state.schedule = None

if "mobile_view" not in st.session_state:
    st.session_state.mobile_view = "schedule"  # schedule, analytics, export

# ---------------- SHIFT CONFIGURATION ----------------
SHIFT_CONFIG = {
    "A": {"name": "Morning", "time": "06:00-14:00", "night": False, "icon": "🌅"},
    "B": {"name": "Afternoon", "time": "14:00-22:00", "night": False, "icon": "☀️"},
    "C": {"name": "Night", "time": "22:00-06:00", "night": True, "icon": "🌙"},
    "G": {"name": "General", "time": "08:00-17:00", "night": False, "icon": "💼"},
    "WO": {"name": "Off/Leave", "time": "Off Duty", "night": False, "icon": "🏖️"}
}

# ---------------- MOBILE BOTTOM NAVIGATION ----------------
st.markdown("---")
nav_col1, nav_col2, nav_col3 = st.columns(3)

with nav_col1:
    if st.button("📋 Schedule", use_container_width=True):
        st.session_state.mobile_view = "schedule"
        st.rerun()

with nav_col2:
    if st.button("📊 Analytics", use_container_width=True):
        st.session_state.mobile_view = "analytics"
        st.rerun()

with nav_col3:
    if st.button("📥 Export", use_container_width=True):
        st.session_state.mobile_view = "export"
        st.rerun()

st.markdown("---")

# ---------------- SCHEDULE VIEW ----------------
if st.session_state.mobile_view == "schedule":
    
    # Year/Month Selection (Collapsible)
    with st.expander("📅 Select Month", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            year = st.number_input("Year", 2020, 2100, 2026, label_visibility="collapsed")
        with col2:
            month = st.selectbox("Month", list(calendar.month_name)[1:], label_visibility="collapsed")
        month_num = list(calendar.month_name).index(month)
        days_in_month = calendar.monthrange(year, month_num)[1]
    
    # Employee Management (Collapsible)
    with st.expander("👥 Employee Management", expanded=False):
        st.markdown("#### Add New Employee")
        col1, col2 = st.columns(2)
        with col1:
            name = st.text_input("Name", placeholder="Employee Name", key="emp_name")
        with col2:
            code = st.text_input("ID", placeholder="Employee ID", key="emp_id")
        
        department = st.selectbox("Department", ["Maintenance", "Production", "Quality", "Warehouse", "Other"])
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("➕ Add Employee", use_container_width=True):
                if name and code:
                    emp = {"Name": name, "ID": code, "Department": department}
                    if emp not in st.session_state.employees:
                        st.session_state.employees.append(emp)
                        st.success(f"✅ {name} added")
                        st.rerun()
                    else:
                        st.warning("Already exists")
                else:
                    st.error("Enter Name & ID")
        
        with col2:
            if st.button("🗑️ Clear All", use_container_width=True):
                st.session_state.employees = []
                st.session_state.schedule = None
                st.rerun()
        
        if st.session_state.employees:
            st.markdown("#### Current Employees")
            emp_df = pd.DataFrame(st.session_state.employees)
            st.dataframe(emp_df, use_container_width=True, hide_index=True)
            st.caption(f"Total: {len(st.session_state.employees)} employees")
    
    # Generate Schedule Button
    if st.button("🔄 Generate New Schedule", type="primary", use_container_width=True):
        if st.session_state.employees:
            data = []
            for emp in st.session_state.employees:
                row = {
                    "Employee Name": emp["Name"],
                    "Employee ID": emp["ID"],
                    "Department": emp["Department"]
                }
                for d in range(1, days_in_month + 1):
                    row[str(d)] = ""
                data.append(row)
            
            st.session_state.schedule = pd.DataFrame(data)
            st.success("✅ Schedule generated!")
            st.balloons()
        else:
            st.warning("Add employees first")
    
    # Display and Edit Schedule
    if st.session_state.schedule is not None:
        st.markdown(f"### 📅 {month} {year}")
        
        # Shift quick picker (for faster data entry on mobile)
        with st.expander("⚡ Quick Shift Select", expanded=False):
            st.caption("Tap a shift to copy its code:")
            shift_cols = st.columns(5)
            for idx, (code, config) in enumerate(SHIFT_CONFIG.items()):
                with shift_cols[idx]:
                    if st.button(f"{config['icon']} {code}", key=f"quick_{code}"):
                        st.session_state.quick_shift = code
                        st.toast(f"Selected: {config['name']} ({code})", icon="✅")
        
        column_config = {
            "Employee Name": st.column_config.TextColumn(disabled=True, width="small"),
            "Employee ID": st.column_config.TextColumn(disabled=True, width="small"),
            "Department": st.column_config.TextColumn(disabled=True, width="small"),
        }
        
        # Show fewer columns on mobile (first 7 days + scroll)
        display_cols = ["Employee Name", "Employee ID", "Department"] + [str(d) for d in range(1, min(31, days_in_month + 1))]
        
        st.info("💡 Tip: Swipe horizontally to see more days. Tap a cell to select shift type.")
        
        edited_df = st.data_editor(
            st.session_state.schedule[display_cols],
            use_container_width=True,
            column_config=column_config,
            height=500,
            key="schedule_editor"
        )
        
        # Update full schedule
        for col in display_cols:
            if col in edited_df.columns:
                st.session_state.schedule[col] = edited_df[col]
        
        # Mobile summary cards
        st.markdown("### 📊 Quick Stats")
        stats_cols = st.columns(3)
        
        total_shifts = 0
        night_shifts = 0
        total_leaves = 0
        
        for _, row in st.session_state.schedule.iterrows():
            for col in st.session_state.schedule.columns:
                if col not in ["Employee Name", "Employee ID", "Department"]:
                    shift = row[col]
                    if shift:
                        total_shifts += 1
                        if SHIFT_CONFIG.get(shift, {}).get("night", False):
                            night_shifts += 1
                        if shift == "WO":
                            total_leaves += 1
        
        with stats_cols[0]:
            st.metric("Total Shifts", total_shifts)
        with stats_cols[1]:
            st.metric("Night Shifts", night_shifts)
        with stats_cols[2]:
            st.metric("Leaves", total_leaves)

# ---------------- ANALYTICS VIEW (MOBILE OPTIMIZED) ----------------
elif st.session_state.mobile_view == "analytics":
    st.header("📊 Analytics")
    
    if st.session_state.schedule is not None:
        # Analytics type selector
        analytics_type = st.selectbox(
            "Select Analysis",
            ["Person-wise", "Shift-wise", "Night Shift", "Leaves", "Summary"],
            label_visibility="collapsed"
        )
        
        if analytics_type == "Person-wise":
            st.subheader("👤 Person-wise Trends")
            
            employees_list = st.session_state.schedule["Employee Name"].tolist()
            selected_employee = st.selectbox("Select Employee", employees_list)
            
            if selected_employee:
                emp_data = st.session_state.schedule[st.session_state.schedule["Employee Name"] == selected_employee]
                
                if not emp_data.empty:
                    shift_counts = {}
                    night_shifts = 0
                    leaves = 0
                    
                    for col in st.session_state.schedule.columns:
                        if col not in ["Employee Name", "Employee ID", "Department"]:
                            shift = emp_data[col].values[0] if len(emp_data) > 0 else ""
                            if shift:
                                shift_counts[shift] = shift_counts.get(shift, 0) + 1
                                if SHIFT_CONFIG.get(shift, {}).get("night", False):
                                    night_shifts += 1
                                if shift == "WO":
                                    leaves += 1
                    
                    total_days = len([c for c in st.session_state.schedule.columns if c not in ["Employee Name", "Employee ID", "Department"]])
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total Days", total_days)
                    with col2:
                        st.metric("Night Shifts", night_shifts)
                    with col3:
                        st.metric("Leaves", leaves)
                    
                    if shift_counts:
                        pie_data = []
                        for shift, count in shift_counts.items():
                            pie_data.append({
                                "Shift": SHIFT_CONFIG.get(shift, {}).get("name", shift),
                                "Count": count
                            })
                        pie_df = pd.DataFrame(pie_data)
                        fig = px.pie(pie_df, values="Count", names="Shift", title=f"{selected_employee}")
                        fig.update_layout(height=400)
                        st.plotly_chart(fig, use_container_width=True)
        
        elif analytics_type == "Shift-wise":
            st.subheader("🔄 Shift Distribution")
            
            shift_totals = {}
            for _, row in st.session_state.schedule.iterrows():
                for col in st.session_state.schedule.columns:
                    if col not in ["Employee Name", "Employee ID", "Department"]:
                        shift = row[col]
                        if shift:
                            shift_totals[shift] = shift_totals.get(shift, 0) + 1
            
            if shift_totals:
                shift_data = []
                for shift_code, count in shift_totals.items():
                    shift_data.append({
                        "Shift": SHIFT_CONFIG.get(shift_code, {}).get("name", shift_code),
                        "Code": shift_code,
                        "Count": count
                    })
                shift_df = pd.DataFrame(shift_data)
                fig = px.bar(shift_df, x="Shift", y="Count", title="Shift Distribution")
                fig.update_layout(height=400)
                st.plotly_chart(fig, use_container_width=True)
                
                st.dataframe(shift_df, use_container_width=True, hide_index=True)
        
        elif analytics_type == "Night Shift":
            st.subheader("🌙 Night Shift Analysis")
            
            night_data = []
            for _, row in st.session_state.schedule.iterrows():
                employee = row["Employee Name"]
                night_count = 0
                consecutive = 0
                max_consecutive = 0
                
                for col in st.session_state.schedule.columns:
                    if col not in ["Employee Name", "Employee ID", "Department"]:
                        shift = row[col]
                        if shift == "C":
                            night_count += 1
                            consecutive += 1
                            max_consecutive = max(max_consecutive, consecutive)
                        else:
                            consecutive = 0
                
                night_data.append({"Employee": employee, "Night Shifts": night_count, "Max Consecutive": max_consecutive})
            
            night_df = pd.DataFrame(night_data)
            fig = px.bar(night_df, x="Employee", y="Night Shifts", title="Night Shifts per Employee")
            fig.update_layout(height=400)
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(night_df, use_container_width=True, hide_index=True)
        
        elif analytics_type == "Leaves":
            st.subheader("🏖️ Leave Analysis")
            
            leave_data = []
            for _, row in st.session_state.schedule.iterrows():
                employee = row["Employee Name"]
                leaves = 0
                for col in st.session_state.schedule.columns:
                    if col not in ["Employee Name", "Employee ID", "Department"]:
                        if row[col] == "WO":
                            leaves += 1
                leave_data.append({"Employee": employee, "Leaves": leaves})
            
            leave_df = pd.DataFrame(leave_data)
            fig = px.bar(leave_df, x="Employee", y="Leaves", title="Leaves per Employee")
            fig.update_layout(height=400)
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(leave_df, use_container_width=True, hide_index=True)
        
        elif analytics_type == "Summary":
            st.subheader("📊 Monthly Summary")
            
            total_employees = len(st.session_state.employees)
            
            shift_totals = {}
            night_total = 0
            leave_total = 0
            
            for _, row in st.session_state.schedule.iterrows():
                for col in st.session_state.schedule.columns:
                    if col not in ["Employee Name", "Employee ID", "Department"]:
                        shift = row[col]
                        if shift:
                            shift_totals[shift] = shift_totals.get(shift, 0) + 1
                            if shift == "C":
                                night_total += 1
                            if shift == "WO":
                                leave_total += 1
            
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Total Employees", total_employees)
                st.metric("Night Shifts", night_total)
            with col2:
                st.metric("Total Leaves", leave_total)
                st.metric("Total Assignments", sum(shift_totals.values()))
    else:
        st.info("Generate a schedule first to see analytics")

# ---------------- EXPORT VIEW ----------------
elif st.session_state.mobile_view == "export":
    st.header("📥 Export Data")
    
    if st.session_state.schedule is not None:
        st.info("Export your schedule to Excel for sharing and printing")
        
        # Preview before export
        with st.expander("Preview Schedule", expanded=False):
            st.dataframe(st.session_state.schedule.head(), use_container_width=True)
        
        if st.button("📊 Generate Excel Report", type="primary", use_container_width=True):
            wb = Workbook()
            
            # Sheet 1: Shift Schedule
            ws1 = wb.active
            ws1.title = "Shift Schedule"
            
            df = st.session_state.schedule
            
            # Add title
            ws1.merge_cells('A1:Z1')
            ws1['A1'] = f"Shift Schedule - {month} {year}"
            ws1['A1'].font = Font(size=14, bold=True)
            ws1['A1'].alignment = Alignment(horizontal='center')
            
            # Add headers
            headers = list(df.columns)
            for col_num, header in enumerate(headers, 1):
                cell = ws1.cell(row=3, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            for row_num, row_data in enumerate(df.values, 4):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws1.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center')
            
            # Auto adjust column widths
            for column in ws1.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws1.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 2: Summary
            ws2 = wb.create_sheet("Summary")
            ws2.append(["Metric", "Value"])
            
            # Calculate summary
            total_employees = len(st.session_state.employees)
            total_shifts = 0
            night_shifts = 0
            leaves = 0
            
            for _, row in st.session_state.schedule.iterrows():
                for col in st.session_state.schedule.columns:
                    if col not in ["Employee Name", "Employee ID", "Department"]:
                        shift = row[col]
                        if shift:
                            total_shifts += 1
                            if shift == "C":
                                night_shifts += 1
                            if shift == "WO":
                                leaves += 1
            
            ws2.append(["Total Employees", total_employees])
            ws2.append(["Total Shift Assignments", total_shifts])
            ws2.append(["Night Shifts", night_shifts])
            ws2.append(["Total Leaves", leaves])
            
            # Save
            output_file = "shift_schedule.xlsx"
            wb.save(output_file)
            st.success("✅ Excel file ready!")
            
            with open(output_file, "rb") as f:
                st.download_button(
                    label="📥 Download Excel File",
                    data=f,
                    file_name=f"shift_schedule_{year}_{month}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
    else:
        st.warning("Generate a schedule first")

# ---------------- SHIFT LEGEND (Always visible) ----------------
st.markdown("---")
st.markdown("### ℹ️ Shift Legend")

legend_cols = st.columns(5)
for idx, (code, config) in enumerate(SHIFT_CONFIG.items()):
    with legend_cols[idx]:
        st.markdown(f"""
        <div class="shift-legend">
            <strong>{config['icon']} {code}</strong><br>
            <small>{config['name']}</small>
        </div>
        """, unsafe_allow_html=True)

# ---------------- FOOTER ----------------
st.caption("📱 Swipe horizontally to see all days | Tap cells to edit shifts")
